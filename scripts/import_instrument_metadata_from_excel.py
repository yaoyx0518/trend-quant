from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
SRC_DIR = PROJECT_ROOT / "src"
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from data.storage.db import init_db

DEFAULT_WORKBOOK = (
    PROJECT_ROOT
    / "outputs"
    / "etf_classification"
    / "中国境内ETF全量梳理_分类建议版_已排序.xlsx"
)
DEFAULT_SHEET = "优选ETF_分类建议"
CATEGORY_NAME_ALIASES = {
    "\u5bbd\u57fa\u6307\u6570": "\u5bbd\u57fa",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Import ETF category metadata from the classification workbook into SQLite."
    )
    parser.add_argument("--workbook", default=str(DEFAULT_WORKBOOK))
    parser.add_argument("--sheet", default=DEFAULT_SHEET)
    return parser.parse_args()


def clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in {"nan", "none", "null"}:
        return ""
    return text


def normalize_symbol(value: Any) -> str:
    text = clean(value).upper()
    if "." not in text:
        digits = "".join(ch for ch in text if ch.isdigit())
        if len(digits) == 6:
            return f"{digits}.SS" if digits.startswith(("5", "6")) else f"{digits}.SZ"
        return text
    code, suffix = text.split(".", 1)
    if suffix == "SH":
        suffix = "SS"
    return f"{code}.{suffix}"


def int_or_none(value: Any) -> int | None:
    text = clean(value)
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def factor_tags(value: Any) -> list[str]:
    text = clean(value)
    if not text:
        return []
    return [part.strip() for part in text.split("/") if part.strip()]


def normalize_category_name(value: Any) -> str:
    text = clean(value)
    return CATEGORY_NAME_ALIASES.get(text, text)


def read_rows(workbook_path: Path, sheet_name: str) -> list[dict[str, Any]]:
    if not workbook_path.exists():
        raise FileNotFoundError(f"workbook not found: {workbook_path}")

    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"sheet not found: {sheet_name}")

    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    headers: list[str] = []
    for row in rows:
        candidate = [clean(cell) for cell in row]
        if "ETF代码" in candidate and "建议一级分类" in candidate:
            headers = candidate
            break
    if not headers:
        raise ValueError("header row not found")
    header_idx = {name: idx for idx, name in enumerate(headers) if name}

    required = ["ETF代码", "ETF名称", "建议一级分类", "建议二级分类", "建议三级分类"]
    missing = [name for name in required if name not in header_idx]
    if missing:
        raise ValueError(f"missing required columns: {', '.join(missing)}")

    l1_order: dict[str, int] = {}
    l2_order: dict[tuple[str, str], int] = {}
    l3_order: dict[tuple[str, str, str], int] = {}
    items: list[dict[str, Any]] = []

    for row in rows:
        cell = lambda name: row[header_idx[name]] if name in header_idx and header_idx[name] < len(row) else ""
        symbol = normalize_symbol(cell("ETF代码"))
        if not symbol:
            continue

        l1 = normalize_category_name(cell("建议一级分类"))
        l2 = clean(cell("建议二级分类"))
        l3 = clean(cell("建议三级分类"))
        if l1 and l1 not in l1_order:
            l1_order[l1] = len(l1_order) + 1
        if l1 and l2 and (l1, l2) not in l2_order:
            l2_order[(l1, l2)] = len([key for key in l2_order if key[0] == l1]) + 1
        if l1 and l2 and l3 and (l1, l2, l3) not in l3_order:
            l3_order[(l1, l2, l3)] = len([key for key in l3_order if key[0] == l1 and key[1] == l2]) + 1

        items.append(
            {
                "symbol": symbol,
                "name": clean(cell("ETF名称")),
                "category_l1": l1,
                "category_l2": l2,
                "category_l3": l3,
                "factor_tags": factor_tags(cell("因子标签")) if "因子标签" in header_idx else [],
                "region_tag": clean(cell("地域标签")) if "地域标签" in header_idx else "",
                "sort_order": int_or_none(cell("序号")) if "序号" in header_idx else len(items) + 1,
                "source": workbook_path.name,
            }
        )

    for item in items:
        l1 = item["category_l1"]
        l2 = item["category_l2"]
        l3 = item["category_l3"]
        item["priority_l1"] = l1_order.get(l1)
        item["priority_l2"] = l2_order.get((l1, l2))
        item["priority_l3"] = l3_order.get((l1, l2, l3)) if l3 else None

    return items


def build_categories(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    categories: dict[str, dict[str, Any]] = {}
    for item in items:
        l1 = item["category_l1"]
        l2 = item["category_l2"]
        l3 = item["category_l3"]
        if l1:
            categories.setdefault(
                l1,
                {
                    "path": l1,
                    "level": 1,
                    "name": l1,
                    "parent_path": "",
                    "priority": item.get("priority_l1"),
                },
            )
        if l1 and l2:
            path = f"{l1}-{l2}"
            categories.setdefault(
                path,
                {
                    "path": path,
                    "level": 2,
                    "name": l2,
                    "parent_path": l1,
                    "priority": item.get("priority_l2"),
                },
            )
        if l1 and l2 and l3:
            parent_path = f"{l1}-{l2}"
            path = f"{parent_path}-{l3}"
            categories.setdefault(
                path,
                {
                    "path": path,
                    "level": 3,
                    "name": l3,
                    "parent_path": parent_path,
                    "priority": item.get("priority_l3"),
                },
            )
    return sorted(
        categories.values(),
        key=lambda item: (
            int(item.get("level") or 0),
            str(item.get("parent_path") or ""),
            int(item.get("priority") or 9999),
            str(item.get("name") or ""),
        ),
    )


def main() -> None:
    args = parse_args()
    workbook_path = Path(args.workbook)
    if not workbook_path.is_absolute():
        workbook_path = PROJECT_ROOT / workbook_path

    items = read_rows(workbook_path, args.sheet)
    categories = build_categories(items)
    db = init_db()
    category_count = db.replace_instrument_categories(categories)
    item_count = db.save_instrument_metadata(items)
    print(
        f"imported {item_count} instrument metadata rows and "
        f"{category_count} category rows from {workbook_path}"
    )


if __name__ == "__main__":
    main()
